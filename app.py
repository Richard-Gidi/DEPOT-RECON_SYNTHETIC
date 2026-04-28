import streamlit as st
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
import pdfplumber
import requests as _requests
import os
from datetime import datetime, timedelta

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass


def _cfg(key: str, default: str = "") -> str:
    try:
        return str(st.secrets[key])
    except (KeyError, AttributeError, FileNotFoundError):
        return os.environ.get(key, default)


st.set_page_config(
    page_title="OilCorp | Intelligence Suite",
    page_icon="⛽",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;500;600;700;800&family=DM+Mono:wght@300;400;500&family=DM+Sans:ital,wght@0,300;0,400;0,500;0,600;1,300&display=swap');

:root {
    --ink:          #0B0D11;
    --ink-soft:     #141720;
    --ink-muted:    #1E2330;
    --surface:      #181C25;
    --surface-hi:   #1F2433;
    --line:         rgba(255,255,255,0.07);
    --line-hi:      rgba(255,255,255,0.12);
    --amber:        #F0A500;
    --amber-dim:    #C8850A;
    --amber-glow:   rgba(240,165,0,0.12);
    --amber-text:   #FFD166;
    --cyan:         #00C9B1;
    --cyan-dim:     rgba(0,201,177,0.1);
    --red:          #FF4D4D;
    --red-dim:      rgba(255,77,77,0.1);
    --green:        #3DDB96;
    --green-dim:    rgba(61,219,150,0.1);
    --blue:         #4D9FFF;
    --blue-dim:     rgba(77,159,255,0.1);
    --text-1:       #F4F5F7;
    --text-2:       #8B90A0;
    --text-3:       #4A5060;
    --r-sm:         6px;
    --r-md:         10px;
    --r-lg:         16px;
}

html, body, [class*="css"], .stApp {
    background-color: var(--ink) !important;
    font-family: 'DM Sans', sans-serif !important;
    color: var(--text-1) !important;
}

footer, .stDeployButton { display: none !important; }

.main .block-container {
    padding: 0 2.5rem 5rem 2.5rem !important;
    max-width: 1500px !important;
}

/* ═══ SIDEBAR ═══════════════════════════════════════════════ */
[data-testid="stSidebar"] {
    background: var(--ink-soft) !important;
    border-right: 1px solid var(--line) !important;
    width: 272px !important;
}
[data-testid="stSidebar"] > div:first-child { padding-top: 0 !important; }

[data-testid="stSidebar"] .stRadio > div { gap: 2px !important; }
[data-testid="stSidebar"] .stRadio label {
    cursor: pointer !important;
    border-radius: var(--r-md) !important;
    padding: 0.65rem 1.1rem !important;
    margin: 1px 6px !important;
    transition: all 0.18s cubic-bezier(0.4,0,0.2,1) !important;
    display: flex !important;
    align-items: center !important;
    border: 1px solid transparent !important;
}
[data-testid="stSidebar"] .stRadio label:hover {
    background: var(--surface-hi) !important;
    border-color: var(--line) !important;
}
[data-testid="stSidebar"] .stRadio [data-testid="stMarkdownContainer"] p {
    color: var(--text-2) !important;
    font-size: 0.875rem !important;
    font-weight: 500 !important;
    font-family: 'DM Sans', sans-serif !important;
    letter-spacing: 0.01em !important;
}

/* ═══ PAGE HERO ════════════════════════════════════════════ */
.hero {
    position: relative;
    padding: 3.5rem 3rem 3rem;
    margin: 0 -2.5rem 3rem -2.5rem;
    background: var(--ink-soft);
    border-bottom: 1px solid var(--line);
    overflow: hidden;
}
.hero::before {
    content: '';
    position: absolute;
    top: -1px; left: 0; right: 0;
    height: 2px;
    background: linear-gradient(90deg, transparent 0%, var(--amber) 30%, var(--amber) 70%, transparent 100%);
}
.hero::after {
    content: '';
    position: absolute;
    top: 0; right: 0;
    width: 420px; height: 100%;
    background: radial-gradient(ellipse at 80% 50%, rgba(240,165,0,0.06) 0%, transparent 70%);
    pointer-events: none;
}
.hero-tag {
    display: inline-flex;
    align-items: center;
    gap: 6px;
    font-family: 'DM Mono', monospace;
    font-size: 0.62rem;
    letter-spacing: 0.18em;
    text-transform: uppercase;
    color: var(--amber);
    background: var(--amber-glow);
    border: 1px solid rgba(240,165,0,0.25);
    padding: 0.3rem 0.75rem;
    border-radius: 100px;
    margin-bottom: 1.2rem;
}
.hero-title {
    font-family: 'Syne', sans-serif;
    font-size: 2.6rem;
    font-weight: 800;
    letter-spacing: -0.04em;
    color: var(--text-1);
    line-height: 1.05;
    margin: 0 0 0.6rem 0;
}
.hero-title em {
    font-style: normal;
    color: var(--amber);
}
.hero-sub {
    font-size: 0.95rem;
    color: var(--text-2);
    font-weight: 400;
    max-width: 540px;
    line-height: 1.65;
    margin: 0;
}

/* ═══ SECTION LABELS ═══════════════════════════════════════ */
.sec {
    display: flex;
    align-items: center;
    gap: 1rem;
    margin: 2.5rem 0 1.2rem;
}
.sec-num {
    font-family: 'DM Mono', monospace;
    font-size: 0.6rem;
    letter-spacing: 0.15em;
    color: var(--amber);
    background: var(--amber-glow);
    border: 1px solid rgba(240,165,0,0.2);
    padding: 0.18rem 0.5rem;
    border-radius: 4px;
    white-space: nowrap;
}
.sec-label {
    font-family: 'Syne', sans-serif;
    font-size: 0.7rem;
    font-weight: 700;
    letter-spacing: 0.14em;
    text-transform: uppercase;
    color: var(--text-2);
}
.sec::after {
    content: '';
    flex: 1;
    height: 1px;
    background: var(--line);
}

/* ═══ STAT CARDS ════════════════════════════════════════════ */
.stats {
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 1px;
    background: var(--line);
    border: 1px solid var(--line);
    border-radius: var(--r-lg);
    overflow: hidden;
    margin-bottom: 2.5rem;
}
.stat {
    background: var(--surface);
    padding: 1.5rem 1.75rem;
    position: relative;
}
.stat::before {
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 2px;
}
.stat.a::before { background: var(--amber); }
.stat.b::before { background: var(--cyan); }
.stat.c::before { background: var(--green); }
.stat.d::before { background: var(--blue); }
.stat-icon {
    font-size: 1rem;
    margin-bottom: 0.9rem;
    opacity: 0.6;
}
.stat-val {
    font-family: 'Syne', sans-serif;
    font-size: 2rem;
    font-weight: 800;
    letter-spacing: -0.04em;
    color: var(--text-1);
    line-height: 1;
    margin-bottom: 0.4rem;
}
.stat-label {
    font-family: 'DM Mono', monospace;
    font-size: 0.58rem;
    letter-spacing: 0.14em;
    text-transform: uppercase;
    color: var(--text-3);
    margin-bottom: 0.2rem;
}
.stat-sub {
    font-size: 0.75rem;
    color: var(--text-3);
}

/* ═══ INFO PANEL ════════════════════════════════════════════ */
.info {
    background: rgba(77,159,255,0.05);
    border: 1px solid rgba(77,159,255,0.15);
    border-left: 3px solid var(--blue);
    border-radius: 0 var(--r-md) var(--r-md) 0;
    padding: 1rem 1.4rem;
    font-size: 0.875rem;
    color: var(--text-2);
    margin-bottom: 1.5rem;
    line-height: 1.7;
}
.info strong { color: var(--blue); font-weight: 600; }
.info code {
    background: rgba(77,159,255,0.1);
    padding: 0.1rem 0.45rem;
    border-radius: 4px;
    font-family: 'DM Mono', monospace;
    font-size: 0.78rem;
    color: #93c5fd;
}

/* ═══ CONFIG PANEL ══════════════════════════════════════════ */
.panel {
    background: var(--surface);
    border: 1px solid var(--line);
    border-radius: var(--r-lg);
    padding: 1.75rem;
    margin-bottom: 1.75rem;
}

/* ═══ TABS ══════════════════════════════════════════════════ */
.stTabs [data-baseweb="tab-list"] {
    background: transparent !important;
    border-bottom: 1px solid var(--line) !important;
    gap: 0 !important;
    padding: 0 !important;
}
.stTabs [data-baseweb="tab"] {
    background: transparent !important;
    border: none !important;
    color: var(--text-3) !important;
    font-family: 'DM Sans', sans-serif !important;
    font-weight: 500 !important;
    font-size: 0.875rem !important;
    padding: 0.85rem 1.5rem !important;
    border-bottom: 2px solid transparent !important;
    margin-bottom: -1px !important;
    transition: all 0.15s !important;
    letter-spacing: 0.01em !important;
}
.stTabs [data-baseweb="tab"]:hover { color: var(--text-1) !important; }
.stTabs [aria-selected="true"] {
    color: var(--amber) !important;
    border-bottom-color: var(--amber) !important;
}
.stTabs [data-baseweb="tab-panel"] {
    background: transparent !important;
    padding: 2rem 0 !important;
}

.tab-hd {
    font-family: 'Syne', sans-serif;
    font-size: 1.25rem;
    font-weight: 700;
    letter-spacing: -0.02em;
    color: var(--text-1);
    margin-bottom: 0.3rem;
}
.tab-sub {
    font-size: 0.82rem;
    color: var(--text-3);
    margin-bottom: 1.75rem;
    font-weight: 400;
}

/* ═══ BUTTONS ═══════════════════════════════════════════════ */
.stButton > button {
    background: var(--surface) !important;
    color: var(--text-2) !important;
    border: 1px solid var(--line-hi) !important;
    border-radius: var(--r-md) !important;
    font-family: 'DM Sans', sans-serif !important;
    font-weight: 500 !important;
    font-size: 0.875rem !important;
    padding: 0.6rem 1.4rem !important;
    transition: all 0.15s cubic-bezier(0.4,0,0.2,1) !important;
    letter-spacing: 0.01em !important;
}
.stButton > button:hover {
    background: var(--surface-hi) !important;
    border-color: rgba(255,255,255,0.2) !important;
    color: var(--text-1) !important;
}
.stButton > button[kind="primary"] {
    background: var(--amber) !important;
    border-color: var(--amber) !important;
    color: #0B0D11 !important;
    font-weight: 700 !important;
    letter-spacing: 0.02em !important;
}
.stButton > button[kind="primary"]:hover {
    background: var(--amber-text) !important;
    border-color: var(--amber-text) !important;
    color: #0B0D11 !important;
}
.stDownloadButton > button {
    background: var(--amber-glow) !important;
    color: var(--amber-text) !important;
    border: 1px solid rgba(240,165,0,0.3) !important;
    border-radius: var(--r-md) !important;
    font-family: 'DM Sans', sans-serif !important;
    font-weight: 600 !important;
    width: 100% !important;
    padding: 0.7rem 1.4rem !important;
    transition: all 0.15s !important;
}
.stDownloadButton > button:hover {
    background: rgba(240,165,0,0.18) !important;
    border-color: var(--amber) !important;
}

/* ═══ FORM INPUTS ═══════════════════════════════════════════ */
.stSelectbox > div > div,
.stMultiSelect > div > div {
    background: var(--surface) !important;
    border: 1px solid var(--line-hi) !important;
    border-radius: var(--r-md) !important;
    color: var(--text-1) !important;
}
.stSelectbox label, .stMultiSelect label,
.stTextInput label, .stDateInput label {
    color: var(--text-3) !important;
    font-size: 0.7rem !important;
    font-weight: 500 !important;
    letter-spacing: 0.1em !important;
    text-transform: uppercase !important;
    font-family: 'DM Mono', monospace !important;
    margin-bottom: 4px !important;
}
.stTextInput > div > div > input {
    background: var(--surface) !important;
    border: 1px solid var(--line-hi) !important;
    border-radius: var(--r-md) !important;
    color: var(--text-1) !important;
    font-family: 'DM Sans', sans-serif !important;
}
[data-baseweb="tag"] {
    background: var(--amber-glow) !important;
    border: 1px solid rgba(240,165,0,0.3) !important;
    color: var(--amber-text) !important;
    border-radius: 4px !important;
    font-size: 0.78rem !important;
}

/* ═══ FILE UPLOADER ════════════════════════════════════════ */
[data-testid="stFileUploader"] > div {
    background: var(--surface) !important;
    border: 1.5px dashed var(--line-hi) !important;
    border-radius: var(--r-lg) !important;
    transition: all 0.2s !important;
}
[data-testid="stFileUploader"] > div:hover {
    border-color: var(--amber) !important;
    background: var(--surface-hi) !important;
}

/* ═══ EXPANDERS ════════════════════════════════════════════ */
.stExpander {
    background: var(--surface) !important;
    border: 1px solid var(--line) !important;
    border-radius: var(--r-md) !important;
    margin-bottom: 0.6rem !important;
    overflow: hidden !important;
}
.stExpander > details > summary {
    background: var(--surface) !important;
    color: var(--text-1) !important;
    font-family: 'DM Sans', sans-serif !important;
    font-weight: 600 !important;
    font-size: 0.875rem !important;
    padding: 0.9rem 1.2rem !important;
}
.stExpander > details[open] > summary {
    border-bottom: 1px solid var(--line) !important;
}
.stExpander > details > div {
    background: var(--ink-muted) !important;
    padding: 1.1rem 1.3rem !important;
}

/* ═══ ALERTS ═══════════════════════════════════════════════ */
.stSuccess { background: rgba(61,219,150,0.06) !important; border-left: 3px solid var(--green) !important; border-radius: 0 var(--r-md) var(--r-md) 0 !important; }
.stWarning { background: rgba(240,165,0,0.06) !important;  border-left: 3px solid var(--amber) !important; border-radius: 0 var(--r-md) var(--r-md) 0 !important; }
.stError   { background: rgba(255,77,77,0.06) !important;   border-left: 3px solid var(--red) !important;   border-radius: 0 var(--r-md) var(--r-md) 0 !important; }
.stInfo    { background: rgba(77,159,255,0.06) !important;  border-left: 3px solid var(--blue) !important;  border-radius: 0 var(--r-md) var(--r-md) 0 !important; }

/* ═══ PROGRESS ══════════════════════════════════════════════ */
.stProgress > div > div {
    background: var(--ink-muted) !important;
    border-radius: 100px !important;
    height: 4px !important;
}
.stProgress > div > div > div {
    background: var(--amber) !important;
    border-radius: 100px !important;
}

/* ═══ DATAFRAME ════════════════════════════════════════════ */
[data-testid="stDataFrame"] {
    border-radius: var(--r-md) !important;
    overflow: hidden !important;
}

/* ═══ EXPORT CARD ═══════════════════════════════════════════ */
.xcard {
    background: var(--surface);
    border: 1px solid var(--line);
    border-radius: var(--r-lg);
    padding: 2.5rem;
    text-align: center;
    max-width: 560px;
    margin: 0 auto 2rem;
    position: relative;
    overflow: hidden;
}
.xcard::before {
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 2px;
    background: linear-gradient(90deg, transparent, var(--amber), transparent);
}
.xcard-glyph {
    font-size: 2rem;
    margin-bottom: 1rem;
    display: block;
    filter: grayscale(0.3);
}
.xcard-title {
    font-family: 'Syne', sans-serif;
    font-size: 1.2rem;
    font-weight: 700;
    color: var(--text-1);
    letter-spacing: -0.02em;
    margin-bottom: 0.5rem;
}
.xcard-desc {
    font-size: 0.84rem;
    color: var(--text-2);
    margin-bottom: 1.5rem;
    line-height: 1.65;
}
.badges {
    display: flex;
    gap: 0.4rem;
    flex-wrap: wrap;
    justify-content: center;
    margin-bottom: 1.5rem;
}
.badge {
    background: var(--ink-muted);
    border: 1px solid var(--line);
    border-radius: 100px;
    padding: 0.22rem 0.7rem;
    font-size: 0.68rem;
    color: var(--text-3);
    font-family: 'DM Mono', monospace;
    letter-spacing: 0.05em;
}
.badge.on {
    background: rgba(61,219,150,0.08);
    border-color: rgba(61,219,150,0.25);
    color: var(--green);
}

/* ═══ DEPOT PANEL ═══════════════════════════════════════════ */
.dpanel {
    background: var(--surface);
    border: 1px solid var(--line);
    border-radius: var(--r-lg);
    overflow: hidden;
    margin-bottom: 1rem;
}
.dpanel-hd {
    background: var(--ink-muted);
    border-bottom: 1px solid var(--line);
    padding: 0.85rem 1.3rem;
    display: flex;
    align-items: center;
    justify-content: space-between;
}
.dpanel-title {
    font-family: 'DM Mono', monospace;
    font-size: 0.68rem;
    font-weight: 500;
    letter-spacing: 0.12em;
    text-transform: uppercase;
    color: var(--text-2);
}
.dpanel-count {
    font-family: 'DM Mono', monospace;
    font-size: 0.65rem;
    background: var(--amber-glow);
    color: var(--amber);
    border: 1px solid rgba(240,165,0,0.2);
    border-radius: 100px;
    padding: 0.12rem 0.55rem;
}
.dpanel-body {
    padding: 0.4rem 0;
    max-height: 300px;
    overflow-y: auto;
}
.ditem {
    display: flex;
    align-items: center;
    gap: 0.75rem;
    padding: 0.45rem 1.3rem;
    border-bottom: 1px solid rgba(255,255,255,0.03);
    font-size: 0.82rem;
    color: var(--text-2);
    transition: background 0.1s;
}
.ditem:last-child { border-bottom: none; }
.ditem:hover { background: var(--surface-hi); }
.didx {
    font-family: 'DM Mono', monospace;
    font-size: 0.58rem;
    color: var(--text-3);
    min-width: 18px;
    text-align: right;
}
.dtick { color: var(--green); font-size: 0.75rem; }

/* ═══ API COUNTER ════════════════════════════════════════════ */
.qcount {
    display: flex;
    align-items: center;
    gap: 1rem;
    background: var(--surface);
    border: 1px solid var(--line);
    border-radius: var(--r-md);
    padding: 1rem 1.3rem;
    margin-bottom: 1.25rem;
}
.qnum {
    font-family: 'Syne', sans-serif;
    font-size: 2rem;
    font-weight: 800;
    color: var(--amber);
    line-height: 1;
    letter-spacing: -0.03em;
}
.qtext {
    font-size: 0.82rem;
    color: var(--text-3);
    line-height: 1.5;
}
.qtext strong { color: var(--text-2); font-weight: 600; }

/* ═══ LOG CONSOLE ════════════════════════════════════════════ */
.logcon {
    font-family: 'DM Mono', monospace;
    font-size: 0.7rem;
    background: #050709;
    border: 1px solid var(--line);
    color: #8B90A0;
    padding: 1rem 1.2rem;
    border-radius: var(--r-md);
    max-height: 200px;
    overflow-y: auto;
    line-height: 1.8;
}

/* ═══ EMPTY STATE ════════════════════════════════════════════ */
.empty {
    text-align: center;
    padding: 4.5rem 0;
    color: var(--text-3);
}
.empty-ico { font-size: 2.5rem; margin-bottom: 1rem; opacity: 0.4; }
.empty-txt {
    font-family: 'DM Mono', monospace;
    font-size: 0.65rem;
    letter-spacing: 0.14em;
    text-transform: uppercase;
}

/* ═══ DAILY STATS ════════════════════════════════════════════ */
.dstats {
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 1px;
    background: var(--line);
    border: 1px solid var(--line);
    border-radius: var(--r-lg);
    overflow: hidden;
    margin-bottom: 2.5rem;
}
.dstat {
    background: var(--surface);
    padding: 1.3rem 1.5rem;
}
.dstat-l {
    font-family: 'DM Mono', monospace;
    font-size: 0.58rem;
    letter-spacing: 0.14em;
    text-transform: uppercase;
    color: var(--text-3);
    margin-bottom: 0.5rem;
}
.dstat-v {
    font-family: 'Syne', sans-serif;
    font-size: 1.7rem;
    font-weight: 800;
    letter-spacing: -0.04em;
    color: var(--text-1);
    line-height: 1;
    margin-bottom: 0.25rem;
}
.dstat-s { font-size: 0.72rem; color: var(--text-3); }

/* ═══ DIVIDER ═══════════════════════════════════════════════ */
.div { height: 1px; background: var(--line); margin: 1.5rem 0; }

/* ═══ SCROLLBAR ═════════════════════════════════════════════ */
::-webkit-scrollbar { width: 4px; height: 4px; }
::-webkit-scrollbar-track { background: transparent; }
::-webkit-scrollbar-thumb { background: var(--line-hi); border-radius: 2px; }

/* ═══ DATE INPUT ════════════════════════════════════════════ */
[data-testid="stDateInput"] > div > div {
    background: var(--surface) !important;
    border: 1px solid var(--line-hi) !important;
    border-radius: var(--r-md) !important;
    color: var(--text-1) !important;
}
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════
# EXCEL STYLE HELPERS
# ══════════════════════════════════════════════════════════════
DARK_BLUE  = "1F3864"
MED_BLUE   = "2E75B6"
LIGHT_BLUE = "BDD7EE"
YELLOW     = "FFD966"
ORANGE     = "F4B942"
GREEN      = "E2EFDA"
HEADER_FG  = "FFFFFF"

def _border(style="thin"):
    s = Side(border_style=style, color="000000")
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=11):
    return Font(bold=bold, color=color, size=size, name="Arial")

def _align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ══════════════════════════════════════════════════════════════
# ENVIRONMENT CONFIG
# ══════════════════════════════════════════════════════════════
OILCORP_USER_ID = _cfg("OILCORP_USER_ID", "123293")
OILCORP_BDC_ID  = _cfg("OILCORP_BDC_ID",  "20900")
NPA_COMPANY_ID  = _cfg("NPA_COMPANY_ID",   "1")
NPA_APP_ID      = _cfg("NPA_APP_ID",       "3")

STOCK_TXN_URL = _cfg(
    "NPA_STOCK_TRANSACTION_URL",
    "https://iml.npa-enterprise.com/NewNPA/home/CreateStockTransactionReport",
)
DAILY_ORDERS_URL = _cfg(
    "NPA_DAILY_ORDERS_URL",
    "https://iml.npa-enterprise.com/NewNPA/home/CreateDailyOrderReport",
)

PRODUCT_MAP = {
    "PMS":    int(_cfg("PRODUCT_PREMIUM_ID", "12")),
    "GASOIL": int(_cfg("PRODUCT_GASOIL_ID",  "14")),
    "LPG":    int(_cfg("PRODUCT_LPG_ID",     "28")),
}

MONTHS = {
    1:"January", 2:"February", 3:"March",    4:"April",
    5:"May",     6:"June",     7:"July",      8:"August",
    9:"September",10:"October",11:"November",12:"December"
}


# ══════════════════════════════════════════════════════════════
# DEPOT MAP
# ══════════════════════════════════════════════════════════════
def _load_depot_map() -> dict:
    depot_map = {}
    raw_entries: dict[str, str] = {}

    for key, value in os.environ.items():
        if key.upper().startswith("DEPOT_"):
            raw_entries[key.upper()] = value

    try:
        for key in st.secrets:
            if key.upper().startswith("DEPOT_"):
                raw_entries[key.upper()] = str(st.secrets[key])
    except Exception:
        pass

    fixes = {
        "GHANA OIL COLTD TAKORADI":                  "GHANA OIL CO.LTD, TAKORADI",
        "GOIL LPG BOTTLING PLANT TEMA":              "GOIL LPG BOTTLING PLANT -TEMA",
        "GOIL LPG BOTTLING PLANT KUMASI":            "GOIL LPG BOTTLING PLANT- KUMASI",
        "NEWGAS CYLINDER BOTTLING LIMITED TEMA":     "NEWGAS CYLINDER BOTTLING LIMITED-TEMA",
        "CHASE PETROLEUM TEMA":                      "CHASE PETROLEUM - TEMA",
        "TEMA FUEL COMPANY TFC":                     "TEMA FUEL COMPANY (TFC)",
        "TEMA MULTI PRODUCTS TMPT":                  "TEMA MULTI PRODUCTS (TMPT)",
        "TEMA OIL REFINERY TOR":                     "TEMA OIL REFINERY (TOR)",
        "GHANA OIL COMPANY LTD SEKONDI NAVAL BASE":  "GHANA OIL COMPANY LTD (SEKONDI NAVAL BASE)",
        "GHANSTOCK LIMITED TAKORADI":                "GHANSTOCK LIMITED (TAKORADI)",
        "SENTUO OIL REFINERY TEMA":                  "SENTUO OIL REFINERY - TEMA",
    }

    for key, value in raw_entries.items():
        raw  = key[6:]
        name = raw.replace("_", " ").strip()
        name = fixes.get(name, name)
        try:
            depot_map[name] = int(value)
        except ValueError:
            pass

    return depot_map


DEPOT_MAP = _load_depot_map()

_HTTP_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": "application/pdf,text/html,*/*;q=0.8",
}


# ══════════════════════════════════════════════════════════════
# HTTP / PDF FETCH
# ══════════════════════════════════════════════════════════════
def _fetch_pdf(url: str, params: dict, timeout: int = 90, debug_key: str = "") -> bytes | None:
    try:
        r = _requests.get(url, params=params, headers=_HTTP_HEADERS, timeout=timeout)
        if debug_key:
            st.session_state[debug_key] = {
                "status_code":    r.status_code,
                "content_type":   r.headers.get("Content-Type", ""),
                "content_length": len(r.content),
                "first_bytes":    r.content[:8].hex() if r.content else "",
                "is_pdf":         r.content[:4] == b"%PDF" if r.content else False,
                "final_url":      r.url,
            }
        r.raise_for_status()
        return r.content if r.content[:4] == b"%PDF" else None
    except Exception as exc:
        if debug_key:
            st.session_state[debug_key] = {"error": str(exc), "final_url": ""}
        return None


# ══════════════════════════════════════════════════════════════
# STOCK TRANSACTION PDF PARSER
# ══════════════════════════════════════════════════════════════
DESCRIPTIONS = sorted([
    "Balance b/fwd", "Stock Take", "Sale",
    "Custody Transfer In", "Custody Transfer Out", "Product Outturn",
], key=len, reverse=True)

SKIP_PFX = (
    "national petroleum authority", "stock transaction report",
    "bdc :", "depot :", "product :", "printed by", "printed on",
    "date trans #", "actual stock balance", "stock commitments",
    "available stock balance", "last stock update", "i.t.s from",
)

def _skip(line):
    lo = line.strip().lower()
    return lo.startswith(SKIP_PFX) or bool(re.match(r"^\d{1,2}\s+\w+,\s+\d{4}", line.strip()))

def _pnum(s):
    s = s.strip()
    neg = s.startswith("(") and s.endswith(")")
    try:
        v = int(s.strip("()").replace(",", ""))
        return -v if neg else v
    except ValueError:
        return None

_DATE_LINE_RE_STOCK = re.compile(r"^(\d{2}/\d{2}/\d{4})\s+(\S+)\s+(.*)")
_TAIL_RE = re.compile(r"(\([\d,]+\)|[\d,]+)\s+(\([\d,]+\)|[\d,]+)\s*$")

def _parse_any_date_line(line: str) -> dict | None:
    line = line.strip()
    m = _DATE_LINE_RE_STOCK.match(line)
    if not m:
        return None
    date  = m.group(1)
    trans = m.group(2)
    rest  = m.group(3).strip()
    for search_str in (rest, trans + " " + rest):
        tail = _TAIL_RE.search(search_str)
        if not tail:
            continue
        vol = _pnum(tail.group(1))
        bal = _pnum(tail.group(2))
        if vol is None or bal is None:
            continue
        middle = search_str[: tail.start()].strip()
        for d in DESCRIPTIONS:
            if middle.lower().startswith(d.lower()):
                acct       = middle[len(d):].strip()
                real_trans = "" if d == "Balance b/fwd" else trans
                return {
                    "Date":        date,
                    "Trans #":     real_trans,
                    "Description": d,
                    "Account":     acct,
                    "Volume":      vol,
                    "Balance":     bal,
                }
    return None

def parse_stock_transaction_pdf(pdf_bytes: bytes) -> list:
    records = []
    seen    = set()
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if not text:
                    continue
                for raw in text.split("\n"):
                    line = raw.strip()
                    if not line or _skip(line):
                        continue
                    row = _parse_any_date_line(line)
                    if row:
                        key = (row["Date"], row["Trans #"], row["Description"], row["Volume"])
                        if key not in seen:
                            seen.add(key)
                            records.append(row)
    except Exception:
        pass
    return records


# ══════════════════════════════════════════════════════════════
# STOCK BALANCE FETCHER
# ══════════════════════════════════════════════════════════════
def fetch_oilcorp_stock_balances(year, month, depot_name, depot_id, product, product_id):
    import calendar
    last_day  = calendar.monthrange(year, month)[1]
    start_str = f"{month:02d}/01/{year}"
    end_str   = f"{month:02d}/{last_day:02d}/{year}"
    params = {
        "lngProductId": product_id,
        "lngBDCId":     OILCORP_BDC_ID,
        "lngDepotId":   depot_id,
        "dtpStartDate": start_str,
        "dtpEndDate":   end_str,
        "lngUserId":    OILCORP_USER_ID,
    }
    pdf_bytes = _fetch_pdf(STOCK_TXN_URL, params)
    if not pdf_bytes:
        return {"opening": None, "closing": None, "records": [], "error": "No PDF returned"}
    records = parse_stock_transaction_pdf(pdf_bytes)
    if not records:
        return {"opening": None, "closing": None, "records": [], "error": "No transactions parsed"}

    bfwd_balance = None
    bfwd_index   = None
    first_stock_take = None
    for i, r in enumerate(records):
        if r["Description"] == "Balance b/fwd" and bfwd_balance is None:
            bfwd_balance = float(r["Balance"])
            bfwd_index   = i
        elif (r["Description"] == "Stock Take" and bfwd_index is not None
              and i > bfwd_index and first_stock_take is None):
            first_stock_take = float(r["Balance"])

    opening = first_stock_take if first_stock_take is not None else bfwd_balance
    closing = None
    for r in reversed(records):
        if r["Description"] != "Balance b/fwd":
            closing = float(r["Balance"])
            break
    if closing is None and records:
        closing = float(records[-1]["Balance"])

    return {"opening": opening, "closing": closing, "records": records, "error": None}


# ══════════════════════════════════════════════════════════════
# DAILY ORDERS PARSER
# ══════════════════════════════════════════════════════════════
_DAILY_DATE_RE   = re.compile(r'^(\d{2}/\d{2}/\d{4})\s+(\S+)\s+(.*)')
_DAILY_STATUS_RE = re.compile(r'^order\s+status\s*:\s*(.+)', re.IGNORECASE)
_DAILY_DEPOT_RE  = re.compile(r'^depot\s*:\s*(.+)',          re.IGNORECASE)

_PRODUCT_PATTERNS = [
    (re.compile(r'AGO\s*\(.*?\)',     re.IGNORECASE), 'AGO'),
    (re.compile(r'PMS\s*\(.*?\)',     re.IGNORECASE), 'PMS'),
    (re.compile(r'GASOIL\s*\(.*?\)', re.IGNORECASE), 'GASOIL'),
    (re.compile(r'\bLPG\b',           re.IGNORECASE), 'LPG'),
    (re.compile(r'\bAGO\b',           re.IGNORECASE), 'AGO'),
    (re.compile(r'\bPMS\b',           re.IGNORECASE), 'PMS'),
    (re.compile(r'\bGASOIL\b',        re.IGNORECASE), 'GASOIL'),
]

_DAILY_SKIP_PREFIXES = (
    "national petroleum authority",
    "daily order report",
    "order date",
    "bdc:",
    "bdc :",
    "printed by",
    "i.t.s from",
    "page ",
)


def _canonicalise_product(raw: str) -> str:
    for pattern, name in _PRODUCT_PATTERNS:
        if pattern.search(raw):
            return name
    return raw.strip()


def _parse_daily_date(tok: str) -> str:
    for fmt in ("%d/%m/%Y", "%d-%b-%Y", "%d-%B-%Y"):
        try:
            return datetime.strptime(tok, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return tok


def extract_oilcorp_daily_orders(pdf_bytes: bytes) -> pd.DataFrame:
    rows       = []
    cur_depot  = "Unknown"
    cur_status = "Unknown"

    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages:
                text = page.extract_text(x_tolerance=3, y_tolerance=3)
                if not text:
                    continue
                for raw_line in text.split("\n"):
                    cl = raw_line.strip()
                    if not cl:
                        continue
                    cl_lower = cl.lower()
                    if any(cl_lower.startswith(p) for p in _DAILY_SKIP_PREFIXES):
                        continue
                    dm = _DAILY_DEPOT_RE.match(cl)
                    if dm:
                        cur_depot  = dm.group(1).strip()
                        cur_status = "Unknown"
                        continue
                    sm = _DAILY_STATUS_RE.match(cl)
                    if sm:
                        cur_status = sm.group(1).strip()
                        continue
                    dlm = _DAILY_DATE_RE.match(cl)
                    if not dlm:
                        continue
                    date_str  = _parse_daily_date(dlm.group(1))
                    order_num = dlm.group(2).strip()
                    rest      = dlm.group(3).strip()
                    tokens    = rest.split()
                    if len(tokens) < 4:
                        continue
                    try:
                        volume = float(tokens[-1].replace(",", ""))
                        price  = float(tokens[-2].replace(",", ""))
                    except ValueError:
                        continue
                    brv         = tokens[-3]
                    product_raw = " ".join(tokens[:-3]).strip()
                    product     = _canonicalise_product(product_raw)
                    rows.append({
                        "Date":         date_str,
                        "Order Number": order_num,
                        "BRV":          brv,
                        "Product":      product,
                        "Depot":        cur_depot,
                        "Quantity (L)": volume,
                        "Price (₵/L)":  price,
                        "Status":       cur_status,
                    })
    except Exception:
        pass

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)
    df = df.drop_duplicates(subset=["Order Number", "BRV", "Date", "Product"])
    try:
        df["_ds"] = pd.to_datetime(df["Date"], errors="coerce")
        df = df.sort_values(["_ds", "Depot", "Product"]).drop(columns=["_ds"])
    except Exception:
        pass
    return df.reset_index(drop=True)


def fetch_oilcorp_daily_orders(start_date: datetime, end_date: datetime) -> pd.DataFrame:
    start_str = start_date.strftime("%m/%d/%Y")
    end_str   = end_date.strftime("%m/%d/%Y")
    params = {
        "lngCompanyId":    NPA_COMPANY_ID,
        "szITSfromPersol": "persol",
        "strGroupBy":      "DEPOT",
        "strGroupBy1":     "",
        "strQuery1":       "",
        "strQuery2":       start_str,
        "strQuery3":       end_str,
        "strQuery4":       "",
        "strPicHeight":    "1",
        "strPicWeight":    "1",
        "intPeriodID":     "-1",
        "iUserId":         OILCORP_USER_ID,
        "iAppId":          NPA_APP_ID,
    }
    pdf_bytes = _fetch_pdf(DAILY_ORDERS_URL, params, debug_key="daily_debug_variant_A")
    if pdf_bytes:
        df = extract_oilcorp_daily_orders(pdf_bytes)
        st.session_state["daily_winning_variant"] = "A"
        return df
    return pd.DataFrame()


# ══════════════════════════════════════════════════════════════
# ORDER REQUEST DATA LOADER + BUILDERS
# ══════════════════════════════════════════════════════════════
def load_order_data(file):
    df = pd.read_excel(file, sheet_name="ORDER REQUEST", header=8)
    df = df[["DATE", "Name of OMC", "Product", "Depot", "Quantity", "Comments"]].copy()
    df.columns = ["DATE", "OMC", "Product", "Depot", "Quantity", "Comments"]
    df = df.dropna(subset=["DATE", "OMC", "Depot", "Quantity"])
    df["DATE"] = pd.to_datetime(df["DATE"], errors="coerce")
    df = df.dropna(subset=["DATE"])
    df["Quantity"] = pd.to_numeric(df["Quantity"], errors="coerce").fillna(0)
    return df


def build_p1_tables(df):
    tables = []
    depots = sorted(df["Depot"].dropna().unique())
    for depot in depots:
        for window, (d_lo, d_hi) in [("W1", (1, 15)), ("W2", (16, 31))]:
            mask = (df["Depot"] == depot) & (df["DATE"].dt.day >= d_lo) & (df["DATE"].dt.day <= d_hi)
            sub = df[mask]
            if sub.empty:
                continue
            for product in sorted(sub["Product"].dropna().unique()):
                psub = sub[sub["Product"] == product]
                if psub.empty:
                    continue
                pivot = psub.groupby("OMC")["Quantity"].sum().reset_index()
                pivot = pivot.sort_values("OMC")
                pivot.columns = ["OMC", "Quantity"]
                pivot["TOTAL"] = pivot["Quantity"]
                tables.append({
                    "title": f"{depot} {product} {window}",
                    "depot": depot, "product": product, "window": window,
                    "data": pivot,
                })
    return tables


def build_summary(df):
    df2 = df.copy()
    df2["DepotGroup"] = df2["Depot"].apply(
        lambda x: "BOST" if str(x).upper().startswith("BOST") else x
    )
    pivot = df2.groupby(["DepotGroup", "Product"])["Quantity"].sum().unstack(fill_value=0)
    for col in ["AGO", "PMS", "LPG"]:
        if col not in pivot.columns:
            pivot[col] = 0
    pivot = pivot[["AGO", "PMS", "LPG"]]
    pivot["GRAND TOTAL"] = pivot.sum(axis=1)
    pivot.loc["GRAND TOTAL"] = pivot.sum()
    pivot = pivot.reset_index()
    pivot.columns.name = None
    return pivot


# ══════════════════════════════════════════════════════════════
# EXCEL WRITERS
# ══════════════════════════════════════════════════════════════
def write_p1_sheet(ws, tables):
    row = 1
    for tbl in tables:
        data  = tbl["data"]
        title = tbl["title"]
        ws.cell(row, 1, title).font = _font(bold=True, color=HEADER_FG, size=12)
        ws.cell(row, 1).fill = _fill(DARK_BLUE)
        ws.cell(row, 1).alignment = _align()
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1
        for ci, label in enumerate(["OMC", "QUANTITY", "TOTAL"], 1):
            c = ws.cell(row, ci, label)
            c.font = _font(bold=True, color=HEADER_FG)
            c.fill = _fill(MED_BLUE)
            c.alignment = _align()
            c.border = _border()
        row += 1
        start_data = row
        for _, r in data.iterrows():
            ws.cell(row, 1, r["OMC"]).alignment = _align(h="left")
            ws.cell(row, 1).border = _border()
            for ci, val in enumerate([r["Quantity"], r["TOTAL"]], 2):
                c = ws.cell(row, ci, val)
                c.number_format = "#,##0"
                c.alignment = _align()
                c.border = _border()
            row += 1
        gt_cell = ws.cell(row, 1, "GRAND TOTAL")
        gt_cell.font = _font(bold=True)
        gt_cell.fill = _fill(YELLOW)
        gt_cell.border = _border()
        gt_cell.alignment = _align(h="left")
        for ci in [2, 3]:
            col_letter = get_column_letter(ci)
            c = ws.cell(row, ci, f"=SUM({col_letter}{start_data}:{col_letter}{row-1})")
            c.font = _font(bold=True)
            c.fill = _fill(YELLOW)
            c.number_format = "#,##0"
            c.alignment = _align()
            c.border = _border()
        row += 3
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18


def write_summary_sheet(ws, summary_df):
    ws.cell(1, 1, "LOADING SUMMARY").font = _font(bold=True, color=HEADER_FG, size=14)
    ws.cell(1, 1).fill = _fill(DARK_BLUE)
    ws.cell(1, 1).alignment = _align()
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    headers = ["DEPOT", "AGO", "PMS", "LPG", "GRAND TOTAL"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(2, ci, h)
        c.font = _font(bold=True, color=HEADER_FG)
        c.fill = _fill(MED_BLUE)
        c.alignment = _align()
        c.border = _border()
    for ri, row_data in summary_df.iterrows():
        excel_row = ri + 3
        is_total  = str(row_data.iloc[0]) == "GRAND TOTAL"
        fill      = _fill(ORANGE) if is_total else _fill(GREEN) if ri % 2 == 0 else PatternFill()
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(excel_row, ci, val)
            c.border = _border()
            c.alignment = _align(h="left" if ci == 1 else "center")
            c.fill = fill
            if ci > 1:
                c.number_format = "#,##0"
            if is_total:
                c.font = _font(bold=True)
    for col, width in zip(["A","B","C","D","E"], [20,14,14,10,16]):
        ws.column_dimensions[col].width = width


def _cell(ws, row, col, value, bold=False, bg=None, fg="000000",
          size=11, h_align="center", num_fmt=None, border=True):
    c = ws.cell(row, col, value)
    c.font = _font(bold=bold, color=fg, size=size)
    c.alignment = _align(h=h_align)
    if bg:
        c.fill = _fill(bg)
    if border:
        c.border = _border()
    if num_fmt:
        c.number_format = num_fmt
    return c


def write_stock_balance_sheet(ws, balance_data, sheet_type, month_label):
    PRODUCTS   = ["PMS", "GASOIL", "LPG"]
    title_text = f"OILCORP ENERGIA LIMITED — {sheet_type} STOCK BALANCE ({month_label})"
    num_cols   = len(PRODUCTS) + 3
    ws.cell(1, 1, title_text).font = _font(bold=True, color=HEADER_FG, size=14)
    ws.cell(1, 1).fill = _fill(DARK_BLUE)
    ws.cell(1, 1).alignment = _align()
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    subtitle = (
        "Balance carried forward at start of month (Stock Take if available, else b/fwd)"
        if sheet_type == "OPENING"
        else "Last recorded running balance at end of month"
    )
    ws.cell(2, 1, subtitle).font = _font(bold=False, color="595959", size=10)
    ws.cell(2, 1).fill = _fill(LIGHT_BLUE)
    ws.cell(2, 1).alignment = _align()
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    headers = ["DEPOT"] + [f"{p} (LT/KG)" for p in PRODUCTS] + ["GRAND TOTAL (LT/KG)", "STATUS"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(3, ci, h)
        c.font = _font(bold=True, color=HEADER_FG, size=11)
        c.fill = _fill(MED_BLUE)
        c.alignment = _align()
        c.border = _border()
    seen_depots = []
    depot_data: dict = {}
    for entry in balance_data:
        depot   = entry["depot"]
        product = entry["product"]
        if depot not in depot_data:
            depot_data[depot] = {}
            seen_depots.append(depot)
        depot_data[depot][product] = {"balance": entry["balance"], "error": entry["error"]}
    row = 4
    product_totals = {p: 0.0 for p in PRODUCTS}
    for depot in seen_depots:
        alt_fill  = "F2F2F2" if row % 2 == 0 else None
        prod_info = depot_data[depot]
        _cell(ws, row, 1, depot, h_align="left", bg=alt_fill, border=True)
        row_total = 0.0
        has_error = False
        error_msgs = []
        for ci, product in enumerate(PRODUCTS, 2):
            info    = prod_info.get(product, {})
            balance = info.get("balance")
            error   = info.get("error")
            if balance is not None:
                _cell(ws, row, ci, balance, num_fmt="#,##0", bg=alt_fill, border=True)
                row_total += balance
                product_totals[product] += balance
            else:
                c = ws.cell(row, ci, "N/A")
                c.font = _font(bold=False, color="FF0000")
                c.alignment = _align()
                c.border = _border()
                if alt_fill:
                    c.fill = _fill(alt_fill)
                has_error = True
                if error:
                    error_msgs.append(f"{product}: {error}")
        total_col = len(PRODUCTS) + 2
        if not has_error or row_total > 0:
            _cell(ws, row, total_col, row_total, bold=True, num_fmt="#,##0", bg=alt_fill, border=True)
        else:
            c = ws.cell(row, total_col, "N/A")
            c.font = _font(bold=False, color="FF0000")
            c.alignment = _align()
            c.border = _border()
        status_col = len(PRODUCTS) + 3
        if not has_error:
            status_txt, status_clr = "✓ OK", "00AA00"
        elif row_total > 0:
            status_txt, status_clr = f"⚠ Partial ({'; '.join(error_msgs)})", "B8860B"
        else:
            status_txt, status_clr = f"✗ {'; '.join(error_msgs) or 'No data'}", "CC0000"
        c = ws.cell(row, status_col, status_txt)
        c.font = _font(bold=False, color=status_clr, size=10)
        c.alignment = _align(h="left")
        c.border = _border()
        if alt_fill:
            c.fill = _fill(alt_fill)
        row += 1
    row += 1
    ws.cell(row, 1, "PRODUCT TOTALS").font = _font(bold=True, color=HEADER_FG, size=11)
    ws.cell(row, 1).fill = _fill(DARK_BLUE)
    ws.cell(row, 1).alignment = _align(h="left")
    ws.cell(row, 1).border = _border()
    grand_total = 0.0
    for ci, product in enumerate(PRODUCTS, 2):
        total = product_totals[product]
        grand_total += total
        c = ws.cell(row, ci, total)
        c.font = _font(bold=True, color=HEADER_FG, size=11)
        c.fill = _fill(DARK_BLUE)
        c.number_format = "#,##0"
        c.alignment = _align()
        c.border = _border()
    total_col = len(PRODUCTS) + 2
    c = ws.cell(row, total_col, grand_total)
    c.font = _font(bold=True, color=HEADER_FG, size=12)
    c.fill = _fill(DARK_BLUE)
    c.number_format = "#,##0"
    c.alignment = _align()
    c.border = _border()
    ws.cell(row, len(PRODUCTS) + 3, "").fill = _fill(DARK_BLUE)
    ws.cell(row, len(PRODUCTS) + 3, "").border = _border()
    ws.column_dimensions["A"].width = 42
    for ci, _ in enumerate(PRODUCTS, 2):
        ws.column_dimensions[get_column_letter(ci)].width = 18
    ws.column_dimensions[get_column_letter(total_col)].width = 22
    ws.column_dimensions[get_column_letter(len(PRODUCTS) + 3)].width = 38


def write_daily_orders_sheet(ws, daily_df: pd.DataFrame):
    if daily_df.empty:
        ws.cell(1, 1, "No daily order data available.")
        return
    title = "OILCORP ENERGIA LIMITED — DAILY ORDERS"
    ws.cell(1, 1, title).font = _font(bold=True, color=HEADER_FG, size=14)
    ws.cell(1, 1).fill = _fill(DARK_BLUE)
    ws.cell(1, 1).alignment = _align()
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(daily_df.columns))
    for ci, col in enumerate(daily_df.columns, 1):
        c = ws.cell(2, ci, col)
        c.font = _font(bold=True, color=HEADER_FG)
        c.fill = _fill(MED_BLUE)
        c.alignment = _align()
        c.border = _border()
    for ri, row_data in daily_df.iterrows():
        excel_row = ri + 3
        alt_fill  = "F2F2F2" if ri % 2 == 0 else None
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(excel_row, ci, val)
            c.border = _border()
            col_name = daily_df.columns[ci - 1]
            c.alignment = _align(h="left" if col_name in ("Date","Depot","BRV","Product","Status") else "center")
            if alt_fill:
                c.fill = _fill(alt_fill)
            if isinstance(val, (int, float)) and col_name in ("Quantity (L)", "Price (₵/L)"):
                c.number_format = "#,##0.00"
    if "Quantity (L)" in daily_df.columns:
        qty_col_idx = list(daily_df.columns).index("Quantity (L)") + 1
        total_row   = len(daily_df) + 3
        c = ws.cell(total_row, 1, "GRAND TOTAL")
        c.font = _font(bold=True)
        c.fill = _fill(YELLOW)
        c.border = _border()
        c.alignment = _align(h="left")
        for ci in range(2, qty_col_idx):
            c = ws.cell(total_row, ci, "")
            c.fill = _fill(YELLOW)
            c.border = _border()
        col_letter = get_column_letter(qty_col_idx)
        c = ws.cell(total_row, qty_col_idx, f"=SUM({col_letter}3:{col_letter}{total_row-1})")
        c.font = _font(bold=True)
        c.fill = _fill(YELLOW)
        c.number_format = "#,##0.00"
        c.alignment = _align()
        c.border = _border()
        for ci in range(qty_col_idx + 1, len(daily_df.columns) + 1):
            c = ws.cell(total_row, ci, "")
            c.fill = _fill(YELLOW)
            c.border = _border()
    for ci, col in enumerate(daily_df.columns, 1):
        ws.column_dimensions[get_column_letter(ci)].width = max(14, len(str(col)) + 4)


# ══════════════════════════════════════════════════════════════
# EXCEL GENERATORS
# ══════════════════════════════════════════════════════════════
def generate_order_request_excel(tables, summary_df, month_label,
                                 opening_data=None, closing_data=None) -> io.BytesIO:
    wb = Workbook()
    ws_p1 = wb.active
    ws_p1.title = "P1"
    write_p1_sheet(ws_p1, tables)
    ws_sum = wb.create_sheet("SUMMARY")
    write_summary_sheet(ws_sum, summary_df)
    if opening_data:
        ws_open = wb.create_sheet("OPENING STOCK")
        write_stock_balance_sheet(ws_open, opening_data, "OPENING", month_label)
    if closing_data:
        ws_close = wb.create_sheet("CLOSING STOCK")
        write_stock_balance_sheet(ws_close, closing_data, "CLOSING", month_label)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_daily_orders_excel(daily_df: pd.DataFrame,
                                start_date: datetime, end_date: datetime) -> io.BytesIO:
    wb = Workbook()
    ws_all = wb.active
    ws_all.title = "Daily Orders"
    write_daily_orders_sheet(ws_all, daily_df)
    if not daily_df.empty and "Product" in daily_df.columns:
        for product in sorted(daily_df["Product"].dropna().unique()):
            prod_df = daily_df[daily_df["Product"] == product].reset_index(drop=True)
            if prod_df.empty:
                continue
            ws_prod = wb.create_sheet(product[:31])
            write_daily_orders_sheet(ws_prod, prod_df)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════
# DEFAULT DEPOTS
# ══════════════════════════════════════════════════════════════
DEFAULT_DEPOTS = [
    "BOST GLOBAL DEPOT",
    "CHASE PETROLEUM - TEMA",
    "PETROLEUM HUB LIMITED",
    "PETROLEUM WARE HOUSE AND SUPPLIES LIMITED",
    "PLATON OIL GAS GHANA LIMITED",
    "QUANTUM LPG LOGISTICS LIMITED",
    "QUANTUM OIL TERMINAL LIMITED",
    "SENTUO OIL REFINERY - TEMA",
    "TEMA FUEL COMPANY (TFC)",
    "TEMA MULTI PRODUCTS (TMPT)",
    "TEMA OIL REFINERY (TOR)",
    "TEMA OIL TERMINAL PLC",
    "VANA ENERGY LIMITED TEMA",
]

def _init_depot_list():
    if "configured_depots" not in st.session_state:
        st.session_state["configured_depots"] = list(DEFAULT_DEPOTS)


# ══════════════════════════════════════════════════════════════
# REUSABLE COMPONENTS
# ══════════════════════════════════════════════════════════════
def render_hero(eyebrow: str, title: str, title_em: str, subtitle: str):
    st.markdown(f"""
    <div class="hero">
        <div class="hero-tag">⛽ {eyebrow}</div>
        <h1 class="hero-title">{title} <em>{title_em}</em></h1>
        <p class="hero-sub">{subtitle}</p>
    </div>
    """, unsafe_allow_html=True)


def render_section(num: str, label: str):
    st.markdown(f"""
    <div class="sec">
        <span class="sec-num">{num}</span>
        <span class="sec-label">{label}</span>
    </div>
    """, unsafe_allow_html=True)


def render_empty(icon: str, msg: str):
    st.markdown(f"""
    <div class="empty">
        <div class="empty-ico">{icon}</div>
        <div class="empty-txt">{msg}</div>
    </div>
    """, unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════
# PAGE: ORDER REQUEST REPORT
# ══════════════════════════════════════════════════════════════
def page_order_report():
    render_hero(
        eyebrow  = "Order Management",
        title    = "Order Request",
        title_em = "Report",
        subtitle = "Upload an ORDER REQUEST sheet to generate P1 breakdowns, loading summaries, and stock balance reports.",
    )

    render_section("01", "Input File")
    uploaded = st.file_uploader(
        "Drop your Excel file here (.xlsx)", type=["xlsx"],
        label_visibility="visible", key="order_upload",
    )

    if not uploaded:
        render_empty("📂", "Awaiting .xlsx upload")
        return

    with st.spinner("Parsing workbook…"):
        try:
            df = load_order_data(uploaded)
        except Exception as e:
            st.error(f"Could not read ORDER REQUEST sheet: {e}")
            return

    df["Year"]  = df["DATE"].dt.year
    df["Month"] = df["DATE"].dt.month
    years       = sorted(df["Year"].unique(), reverse=True)

    render_section("02", "Reporting Period")
    st.markdown('<div class="panel">', unsafe_allow_html=True)
    col1, col2, _ = st.columns([1, 1, 2])
    with col1:
        sel_year = st.selectbox("Year", years, key="or_year")
    with col2:
        avail_months = sorted(df[df["Year"] == sel_year]["Month"].unique())
        sel_month = st.selectbox("Month", avail_months,
                                 format_func=lambda m: MONTHS[m], key="or_month")
    st.markdown('</div>', unsafe_allow_html=True)

    month_label = f"{MONTHS[sel_month]} {sel_year}"
    filtered    = df[(df["Year"] == sel_year) & (df["Month"] == sel_month)].copy()

    if filtered.empty:
        st.warning("No data found for the selected period.")
        return

    w1 = filtered[filtered["DATE"].dt.day <= 15]
    w2 = filtered[filtered["DATE"].dt.day >= 16]

    st.markdown(f"""
    <div class="stats">
      <div class="stat a">
        <div class="stat-label">Total Records</div>
        <div class="stat-val">{len(filtered):,}</div>
        <div class="stat-sub">W1: {len(w1):,} · W2: {len(w2):,}</div>
      </div>
      <div class="stat b">
        <div class="stat-label">Total Volume (L)</div>
        <div class="stat-val">{filtered["Quantity"].sum():,.0f}</div>
        <div class="stat-sub">{month_label}</div>
      </div>
      <div class="stat c">
        <div class="stat-label">Active OMCs</div>
        <div class="stat-val">{filtered["OMC"].nunique()}</div>
        <div class="stat-sub">Unique companies</div>
      </div>
      <div class="stat d">
        <div class="stat-label">Active Depots</div>
        <div class="stat-val">{filtered["Depot"].nunique()}</div>
        <div class="stat-sub">Unique depots</div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    tables     = build_p1_tables(filtered)
    summary_df = build_summary(filtered)

    render_section("03", "Report Views")

    tab1, tab2, tab3, tab4 = st.tabs([
        "P1 — Depot Breakdown",
        "Loading Summary",
        "Stock Balances",
        "Export",
    ])

    with tab1:
        st.markdown('<div class="tab-hd">P1 — Depot / Product / Window</div>', unsafe_allow_html=True)
        st.markdown(f'<div class="tab-sub">Order requests by depot, product, and fortnightly window · {month_label}</div>', unsafe_allow_html=True)
        if not tables:
            st.info("No data to display for this period.")
        else:
            depots_in_data = sorted(set(t["depot"] for t in tables))
            for depot in depots_in_data:
                with st.expander(f"🏭  {depot}", expanded=True):
                    depot_tables = [t for t in tables if t["depot"] == depot]
                    cols = st.columns(min(len(depot_tables), 2))
                    for i, tbl in enumerate(depot_tables):
                        with cols[i % 2]:
                            icon = {"PMS":"🟡","AGO":"🔵","LPG":"🟢"}.get(tbl["product"],"⚪")
                            st.markdown(f"**{icon} {tbl['title']}**")
                            display = tbl["data"][["OMC", "Quantity"]].copy()
                            display.columns = ["OMC", "Quantity (L)"]
                            display["Quantity (L)"] = display["Quantity (L)"].apply(lambda x: f"{x:,.0f}")
                            total = tbl["data"]["Quantity"].sum()
                            total_row = pd.DataFrame([{"OMC":"GRAND TOTAL","Quantity (L)":f"{total:,.0f}"}])
                            display = pd.concat([display, total_row], ignore_index=True)
                            st.dataframe(display, use_container_width=True, hide_index=True)

    with tab2:
        st.markdown('<div class="tab-hd">Loading Summary</div>', unsafe_allow_html=True)
        st.markdown(f'<div class="tab-sub">Aggregate volumes by depot group and product · {month_label}</div>', unsafe_allow_html=True)
        display_sum = summary_df.copy()
        for col in ["AGO", "PMS", "LPG", "GRAND TOTAL"]:
            if col in display_sum.columns:
                display_sum[col] = display_sum[col].apply(
                    lambda x: f"{int(x):,}" if pd.notna(x) and x != 0 else "—"
                )
        col_rename = {"DepotGroup": "DEPOT"} if "DepotGroup" in display_sum.columns else {}
        st.dataframe(display_sum.rename(columns=col_rename), use_container_width=True, hide_index=True)

    with tab3:
        _init_depot_list()
        st.markdown('<div class="tab-hd">OILCORP Stock Balances</div>', unsafe_allow_html=True)
        st.markdown(f'<div class="tab-sub">NPA stock transaction ledger for OILCORP ENERGIA LIMITED · {month_label}</div>', unsafe_allow_html=True)

        st.markdown("""
        <div class="info">
            <strong>Opening Balance</strong> — First <em>Stock Take</em> after the b/fwd row (if present), otherwise the <em>Balance b/fwd</em> value.<br>
            <strong>Closing Balance</strong> — Last recorded running balance at end of the selected month.
        </div>
        """, unsafe_allow_html=True)

        col_manager, col_controls = st.columns([5, 4], gap="large")

        with col_manager:
            configured   = st.session_state["configured_depots"]
            n_configured = len(configured)
            depot_rows_html = "".join([
                f'<div class="ditem"><span class="didx">{i+1}</span>'
                f'<span class="dtick">✓</span><span>{d}</span></div>'
                for i, d in enumerate(configured)
            ])
            st.markdown(f"""
            <div class="dpanel">
              <div class="dpanel-hd">
                <span class="dpanel-title">Depot List</span>
                <span class="dpanel-count">{n_configured} depots</span>
              </div>
              <div class="dpanel-body">{depot_rows_html}</div>
            </div>
            """, unsafe_allow_html=True)

            with st.expander("✏️  Edit depot list", expanded=False):
                st.markdown("**Remove a depot**")
                if configured:
                    remove_choice = st.selectbox("Select depot to remove", options=configured,
                                                  key="depot_remove_choice", label_visibility="collapsed")
                    if st.button("🗑  Remove selected depot", key="btn_remove_depot"):
                        st.session_state["configured_depots"].remove(remove_choice)
                        st.rerun()
                else:
                    st.caption("No depots configured.")

                st.markdown('<div class="div"></div>', unsafe_allow_html=True)
                st.markdown("**Add from configured depots**")
                addable = [d for d in sorted(DEPOT_MAP.keys()) if d not in st.session_state["configured_depots"]]
                if addable:
                    add_choice = st.selectbox("Choose depot to add", options=addable,
                                               key="depot_add_choice", label_visibility="collapsed")
                    if st.button("➕  Add depot", key="btn_add_depot"):
                        st.session_state["configured_depots"].append(add_choice)
                        st.session_state["configured_depots"].sort()
                        st.rerun()
                else:
                    st.caption("All configured depots are already in the list.")

                st.markdown('<div class="div"></div>', unsafe_allow_html=True)
                st.markdown("**Add custom depot name**")
                custom_name = st.text_input("Type a depot name manually",
                                             placeholder="e.g. GHANA OIL CO.LTD, TAKORADI",
                                             key="depot_custom_input", label_visibility="collapsed")
                if st.button("➕  Add custom depot", key="btn_add_custom") and custom_name.strip():
                    name = custom_name.strip()
                    if name not in st.session_state["configured_depots"]:
                        st.session_state["configured_depots"].append(name)
                        st.session_state["configured_depots"].sort()
                        st.rerun()
                    else:
                        st.warning("That depot is already in the list.")

                st.markdown('<div class="div"></div>', unsafe_allow_html=True)
                if st.button("↺  Reset to defaults", key="btn_reset_depots"):
                    st.session_state["configured_depots"] = list(DEFAULT_DEPOTS)
                    st.rerun()

        with col_controls:
            selected_products = st.multiselect(
                "Products to fetch", ["PMS", "GASOIL", "LPG"],
                default=["PMS", "GASOIL", "LPG"], key="oilcorp_products",
            )
            depots_to_query   = st.session_state["configured_depots"]
            products_to_query = selected_products if selected_products else ["PMS", "GASOIL", "LPG"]
            total_calls       = len(depots_to_query) * len(products_to_query)

            st.markdown(f"""
            <div class="qcount">
                <div class="qnum">{total_calls}</div>
                <div class="qtext">API calls &nbsp;·&nbsp; <strong>{len(depots_to_query)}</strong> depot(s) ×
                <strong>{len(products_to_query)}</strong> product(s) for <strong>{month_label}</strong></div>
            </div>
            """, unsafe_allow_html=True)

            if not depots_to_query:
                st.warning("Add at least one depot before fetching.")
            elif st.button("Fetch Stock Balances", type="primary", use_container_width=True):
                opening_results, closing_results = [], []
                progress  = st.progress(0, text="Initialising…")
                log_box   = st.empty()
                log_lines = []
                call_n    = 0

                for depot_name in depots_to_query:
                    depot_id = DEPOT_MAP.get(depot_name)
                    if not depot_id:
                        for prod in products_to_query:
                            opening_results.append({"depot": depot_name, "product": prod, "balance": None, "error": "Depot ID not found in config"})
                            closing_results.append({"depot": depot_name, "product": prod, "balance": None, "error": "Depot ID not found in config"})
                        call_n += len(products_to_query)
                        log_lines.append(f"⚠  {depot_name} — not in config, skipped")
                        log_box.markdown(f"<div class='logcon'>{'<br>'.join(log_lines[-12:])}</div>", unsafe_allow_html=True)
                        continue

                    for product in products_to_query:
                        product_id = PRODUCT_MAP.get(product)
                        call_n    += 1
                        progress.progress(call_n / total_calls, text=f"Fetching {depot_name} · {product}  ({call_n}/{total_calls})")
                        result = fetch_oilcorp_stock_balances(sel_year, sel_month, depot_name, depot_id, product, product_id)
                        if result["opening"] is not None:
                            log_lines.append(f"✓  {depot_name} [{product}] — Open: {result['opening']:,.0f} | Close: {result['closing']:,.0f}")
                        else:
                            log_lines.append(f"⚠  {depot_name} [{product}] — {result['error']}")
                        log_box.markdown(f"<div class='logcon'>{'<br>'.join(log_lines[-12:])}</div>", unsafe_allow_html=True)
                        opening_results.append({"depot": depot_name, "product": product, "balance": result["opening"], "error": result["error"]})
                        closing_results.append({"depot": depot_name, "product": product, "balance": result["closing"], "error": result["error"]})

                progress.progress(1.0, text="Complete")
                st.session_state["oilcorp_opening"] = opening_results
                st.session_state["oilcorp_closing"] = closing_results
                n_ok = sum(1 for r in opening_results if r["balance"] is not None)
                st.success(f"Fetched {total_calls} combinations — **{n_ok}** with data, **{total_calls - n_ok}** with errors.")

        opening_data = st.session_state.get("oilcorp_opening")
        closing_data = st.session_state.get("oilcorp_closing")

        if opening_data or closing_data:
            PRODUCTS = ["PMS", "GASOIL", "LPG"]

            def _pivot_for_display(data):
                depot_order, depot_map_local = [], {}
                for entry in data:
                    d, p = entry["depot"], entry["product"]
                    if d not in depot_map_local:
                        depot_map_local[d] = {}
                        depot_order.append(d)
                    depot_map_local[d][p] = entry
                rows = []
                for depot in depot_order:
                    prod_info = depot_map_local[depot]
                    row_dict  = {"DEPOT": depot}
                    row_total = 0.0
                    has_error = False
                    for prod in PRODUCTS:
                        info    = prod_info.get(prod, {})
                        balance = info.get("balance")
                        if balance is not None:
                            row_dict[prod] = f"{balance:,.0f}"
                            row_total     += balance
                        else:
                            row_dict[prod] = "N/A"
                            has_error = True
                    row_dict["TOTAL"]  = f"{row_total:,.0f}" if (not has_error or row_total > 0) else "N/A"
                    row_dict["STATUS"] = "✓" if not has_error else "⚠ Partial" if row_total > 0 else "✗"
                    rows.append(row_dict)
                totals = {p: 0.0 for p in PRODUCTS}
                for entry in data:
                    if entry["balance"] is not None:
                        totals[entry["product"]] = totals.get(entry["product"], 0.0) + entry["balance"]
                total_row = {"DEPOT": "GRAND TOTAL"}
                grand = 0.0
                for prod in PRODUCTS:
                    total_row[prod] = f"{totals[prod]:,.0f}"
                    grand += totals[prod]
                total_row["TOTAL"]  = f"{grand:,.0f}"
                total_row["STATUS"] = ""
                rows.append(total_row)
                return pd.DataFrame(rows)

            st.markdown('<div class="div"></div>', unsafe_allow_html=True)
            col_a, col_b = st.columns(2)
            with col_a:
                if opening_data:
                    st.markdown("#### 📂 Opening Stock Balance")
                    st.dataframe(_pivot_for_display(opening_data), use_container_width=True, hide_index=True)
            with col_b:
                if closing_data:
                    st.markdown("#### 📁 Closing Stock Balance")
                    st.dataframe(_pivot_for_display(closing_data), use_container_width=True, hide_index=True)

    with tab4:
        opening_data = st.session_state.get("oilcorp_opening")
        closing_data = st.session_state.get("oilcorp_closing")

        open_cls  = "on" if opening_data else ""
        close_cls = "on" if closing_data else ""

        st.markdown(f"""
        <div class="xcard">
            <span class="xcard-glyph">📊</span>
            <div class="xcard-title">Order Request Report — Excel Workbook</div>
            <div class="xcard-desc">
                P1 breakdown, loading summary, and stock balances for <strong>{month_label}</strong>.
                Each section is compiled into a separate, formatted worksheet.
            </div>
            <div class="badges">
                <span class="badge on">P1</span>
                <span class="badge on">SUMMARY</span>
                <span class="badge {open_cls}">OPENING STOCK</span>
                <span class="badge {close_cls}">CLOSING STOCK</span>
            </div>
        </div>
        """, unsafe_allow_html=True)

        if not (opening_data or closing_data):
            st.markdown("""
            <div class="info" style="max-width:520px; margin:0 auto 1rem auto; text-align:center;">
                Visit the <strong>Stock Balances</strong> tab to fetch OILCORP stock data first.
            </div>
            """, unsafe_allow_html=True)

        col_left, col_mid, col_right = st.columns([1, 2, 1])
        with col_mid:
            if st.button("Build Order Request Report", type="primary", use_container_width=True):
                with st.spinner("Compiling workbook…"):
                    excel_buf = generate_order_request_excel(
                        tables, summary_df, month_label,
                        opening_data=opening_data,
                        closing_data=closing_data,
                    )
                fname = f"OilCorp_OrderRequest_{MONTHS[sel_month]}_{sel_year}.xlsx"
                st.download_button(
                    label=f"⬇  Download  {fname}",
                    data=excel_buf,
                    file_name=fname,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )


# ══════════════════════════════════════════════════════════════
# PAGE: DAILY ORDERS
# ══════════════════════════════════════════════════════════════
def page_daily_orders():
    render_hero(
        eyebrow  = "Dispatch Intelligence",
        title    = "Daily",
        title_em = "Orders",
        subtitle = "Truck-level dispatch records for OilCorp Energia, fetched live from the NPA portal.",
    )

    render_section("01", "Date Range")
    st.markdown('<div class="panel">', unsafe_allow_html=True)
    col1, col2, col3 = st.columns([1, 1, 2])
    with col1:
        start_date = st.date_input("From", value=datetime.now() - timedelta(days=7), key="daily_start")
    with col2:
        end_date   = st.date_input("To",   value=datetime.now(),                     key="daily_end")
    with col3:
        st.markdown("<br>", unsafe_allow_html=True)
        st.info(f"**{(end_date - start_date).days + 1}** day window selected")
    st.markdown('</div>', unsafe_allow_html=True)

    if start_date > end_date:
        st.error("Start date must be before end date.")
        return

    render_section("02", "Fetch Orders")
    st.markdown(f"""
    <div class="info">
        Queries the NPA Daily Order Report endpoint (<code>CreateDailyOrderReport</code>) authenticated as
        <strong>OilCorp Energia</strong> (iUserId=<code>{OILCORP_USER_ID}</code>).
        Results are automatically scoped to OilCorp's own orders — covers all products (PMS, GASOIL, LPG)
        and all depots, grouped by depot.<br><br>
        <strong>Config:</strong> lngCompanyId=<code>{NPA_COMPANY_ID}</code> &nbsp;|&nbsp;
        iUserId=<code>{OILCORP_USER_ID}</code> &nbsp;|&nbsp; iAppId=<code>{NPA_APP_ID}</code>
    </div>
    """, unsafe_allow_html=True)

    col_btn, col_dbg = st.columns([1, 3])
    with col_btn:
        fetch_clicked = st.button("Fetch Daily Orders", type="primary", use_container_width=True)
    with col_dbg:
        show_debug = st.checkbox("Show debug info", value=False, key="daily_show_debug")

    if fetch_clicked:
        for k in ["daily_debug_variant_A", "daily_winning_variant"]:
            st.session_state.pop(k, None)
        with st.spinner(f"Fetching OilCorp daily orders ({start_date.strftime('%d %b')} → {end_date.strftime('%d %b %Y')})…"):
            try:
                df_fetched = fetch_oilcorp_daily_orders(
                    datetime.combine(start_date, datetime.min.time()),
                    datetime.combine(end_date,   datetime.min.time()),
                )
                st.session_state["oilcorp_daily_df"]    = df_fetched
                st.session_state["oilcorp_daily_start"] = start_date
                st.session_state["oilcorp_daily_end"]   = end_date
                if df_fetched.empty:
                    st.warning("No records found for this date range. Check debug info below.")
                else:
                    winning = st.session_state.get("daily_winning_variant", "?")
                    st.success(f"✅ {len(df_fetched):,} order records retrieved (Variant {winning}).")
            except Exception as e:
                st.error(f"Fetch failed: {e}")
                return

    if show_debug:
        render_section("🔍", "Debug — API Response")
        dbg = st.session_state.get("daily_debug_variant_A")
        if dbg:
            with st.expander("Variant A response", expanded=True):
                st.json(dbg)
        else:
            st.caption("Variant A: not yet fetched — click Fetch Daily Orders first.")

    df = st.session_state.get("oilcorp_daily_df", pd.DataFrame())
    if df is None or (isinstance(df, pd.DataFrame) and df.empty):
        render_empty("📋", "Select a date range and fetch to see orders")
        return

    d_start    = st.session_state.get("oilcorp_daily_start", start_date)
    d_end      = st.session_state.get("oilcorp_daily_end",   end_date)
    period_lbl = f"{d_start.strftime('%d %b')} – {d_end.strftime('%d %b %Y')}"

    total_vol = df["Quantity (L)"].sum() if "Quantity (L)" in df.columns else 0
    n_orders  = len(df)
    n_depots  = df["Depot"].nunique() if "Depot" in df.columns else 0
    total_val = (
        (df["Quantity (L)"] * df["Price (₵/L)"]).sum()
        if all(c in df.columns for c in ["Quantity (L)", "Price (₵/L)"])
        else 0
    )

    render_section("03", "Summary")
    st.markdown(f"""
    <div class="dstats">
      <div class="dstat">
        <div class="dstat-l">Total Orders</div>
        <div class="dstat-v">{n_orders:,}</div>
        <div class="dstat-s">{period_lbl}</div>
      </div>
      <div class="dstat">
        <div class="dstat-l">Total Volume</div>
        <div class="dstat-v">{total_vol:,.0f}</div>
        <div class="dstat-s">Litres dispatched</div>
      </div>
      <div class="dstat">
        <div class="dstat-l">Depots Active</div>
        <div class="dstat-v">{n_depots}</div>
        <div class="dstat-s">Unique depots</div>
      </div>
      <div class="dstat">
        <div class="dstat-l">Est. Value</div>
        <div class="dstat-v">₵{total_val:,.0f}</div>
        <div class="dstat-s">Volume × price</div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    render_section("04", "Breakdown")

    tab_prod, tab_depot, tab_detail = st.tabs([
        "By Product", "By Depot", "Full Detail",
    ])

    with tab_prod:
        st.markdown('<div class="tab-hd">Product Summary</div>', unsafe_allow_html=True)
        if "Product" in df.columns and "Quantity (L)" in df.columns:
            prod_grp = (
                df.groupby("Product")
                .agg(Orders=("Order Number","count"), Volume=("Quantity (L)","sum"))
                .reset_index()
                .sort_values("Volume", ascending=False)
            )
            prod_grp["Volume"] = prod_grp["Volume"].apply(lambda x: f"{x:,.0f}")
            st.dataframe(prod_grp, use_container_width=True, hide_index=True)
        else:
            st.info("Product column not available.")

    with tab_depot:
        st.markdown('<div class="tab-hd">Depot Summary</div>', unsafe_allow_html=True)
        if "Depot" in df.columns and "Quantity (L)" in df.columns:
            depot_grp = (
                df.groupby("Depot")
                .agg(Orders=("Order Number","count"), Volume=("Quantity (L)","sum"))
                .reset_index()
                .sort_values("Volume", ascending=False)
            )
            depot_grp["Volume"] = depot_grp["Volume"].apply(lambda x: f"{x:,.0f}")
            st.dataframe(depot_grp, use_container_width=True, hide_index=True)
        else:
            st.info("Depot column not available.")

    with tab_detail:
        st.markdown('<div class="tab-hd">Full Order Detail</div>', unsafe_allow_html=True)
        fc1, fc2, fc3 = st.columns(3)
        with fc1:
            products_all = ["All"] + sorted(df["Product"].dropna().unique().tolist()) if "Product" in df.columns else ["All"]
            prod_filter  = st.selectbox("Filter by Product", products_all, key="daily_prod_filter")
        with fc2:
            depots_all   = ["All"] + sorted(df["Depot"].dropna().unique().tolist()) if "Depot" in df.columns else ["All"]
            depot_filter = st.selectbox("Filter by Depot", depots_all, key="daily_depot_filter")
        with fc3:
            status_all   = ["All"] + sorted(df["Status"].dropna().unique().tolist()) if "Status" in df.columns else ["All"]
            status_filter = st.selectbox("Filter by Status", status_all, key="daily_status_filter")

        filt = df.copy()
        if prod_filter   != "All" and "Product" in filt.columns: filt = filt[filt["Product"]  == prod_filter]
        if depot_filter  != "All" and "Depot"   in filt.columns: filt = filt[filt["Depot"]    == depot_filter]
        if status_filter != "All" and "Status"  in filt.columns: filt = filt[filt["Status"]   == status_filter]

        filt_vol = filt["Quantity (L)"].sum() if "Quantity (L)" in filt.columns else 0
        st.caption(f"Showing **{len(filt):,}** records  |  Volume: **{filt_vol:,.0f} L**")
        st.dataframe(filt, use_container_width=True, hide_index=True, height=450)

    render_section("05", "Export")

    products_in_df = sorted(df["Product"].dropna().unique()) if "Product" in df.columns else []
    badges_html = (
        '<span class="badge on">Daily Orders</span>'
        + "".join(f'<span class="badge on">{p}</span>' for p in products_in_df)
    )

    st.markdown(f"""
    <div class="xcard">
        <span class="xcard-glyph">📦</span>
        <div class="xcard-title">Daily Orders — Excel Workbook</div>
        <div class="xcard-desc">
            All orders plus per-product breakdowns for <strong>{period_lbl}</strong>.<br>
            Columns: Date, Order Number, BRV, Product, Depot, Quantity (L), Price (₵/L), Status.
        </div>
        <div class="badges">{badges_html}</div>
    </div>
    """, unsafe_allow_html=True)

    col_l, col_m, col_r = st.columns([1, 2, 1])
    with col_m:
        if st.button("Build Daily Orders Report", type="primary", use_container_width=True):
            with st.spinner("Compiling workbook…"):
                excel_buf = generate_daily_orders_excel(df, d_start, d_end)
            fname = f"OilCorp_DailyOrders_{d_start.strftime('%Y%m%d')}_{d_end.strftime('%Y%m%d')}.xlsx"
            st.download_button(
                label=f"⬇  Download {fname}",
                data=excel_buf,
                file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )


# ══════════════════════════════════════════════════════════════
# SIDEBAR + MAIN
# ══════════════════════════════════════════════════════════════
def render_sidebar():
    with st.sidebar:
        st.markdown("""
        <div style="padding: 2rem 1.4rem 1.2rem; border-bottom: 1px solid rgba(255,255,255,0.06);">
            <div style="font-family: 'DM Mono', monospace; font-size: 0.55rem; letter-spacing: 0.2em;
                        text-transform: uppercase; color: #F0A500; margin-bottom: 0.6rem;">
                ⛽ OilCorp Energia
            </div>
            <div style="font-family: 'Syne', sans-serif; font-size: 1.4rem; font-weight: 800;
                        color: #F4F5F7; letter-spacing: -0.03em; line-height: 1.1;">
                Intelligence<br><span style="color: #F0A500;">Suite</span>
            </div>
            <div style="font-family: 'DM Mono', monospace; font-size: 0.65rem; color: #4A5060;
                        margin-top: 0.5rem; letter-spacing: 0.04em;">
                NPA Reporting Platform
            </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("""
        <div style="font-family: 'DM Mono', monospace; font-size: 0.52rem; letter-spacing: 0.18em;
                    text-transform: uppercase; color: #4A5060; padding: 1.2rem 1.4rem 0.5rem;">
            Navigation
        </div>
        """, unsafe_allow_html=True)

        page = st.radio(
            "nav",
            options=["Order Request Report", "Daily Orders"],
            label_visibility="collapsed",
            key="nav_page",
        )

        st.markdown('<div style="height: 1px; background: rgba(255,255,255,0.06); margin: 1rem 1.2rem;"></div>', unsafe_allow_html=True)

        st.markdown("""
        <div style="font-family: 'DM Mono', monospace; font-size: 0.52rem; letter-spacing: 0.18em;
                    text-transform: uppercase; color: #4A5060; padding: 0 1.4rem 0.6rem;">
            Session Data
        </div>
        """, unsafe_allow_html=True)

        has_opening = bool(st.session_state.get("oilcorp_opening"))
        has_closing = bool(st.session_state.get("oilcorp_closing"))
        has_daily   = (
            isinstance(st.session_state.get("oilcorp_daily_df"), pd.DataFrame)
            and not st.session_state.get("oilcorp_daily_df", pd.DataFrame()).empty
        )

        def _badge(label, active):
            color = "#3DDB96" if active else "#4A5060"
            sym   = "●" if active else "○"
            return (
                f'<div style="font-family: DM Sans, sans-serif; font-size: 0.8rem; color: {color}; '
                f'padding: 0.2rem 1.4rem; display: flex; align-items: center; gap: 0.4rem;">'
                f'<span style="font-size: 0.6rem;">{sym}</span>{label}</div>'
            )

        st.markdown(
            _badge("Opening Stock", has_opening) +
            _badge("Closing Stock", has_closing) +
            _badge("Daily Orders",  has_daily),
            unsafe_allow_html=True,
        )

        if any([has_opening, has_closing, has_daily]):
            st.markdown('<div style="margin: 0.8rem 1.2rem 0;">', unsafe_allow_html=True)
            if st.button("Clear session data", use_container_width=True):
                for k in ["oilcorp_opening", "oilcorp_closing", "oilcorp_daily_df",
                           "oilcorp_daily_start", "oilcorp_daily_end"]:
                    st.session_state.pop(k, None)
                st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)

        st.markdown("""
        <div style="position: absolute; bottom: 1.5rem; left: 0; right: 0; padding: 0 1.4rem; text-align: center;">
            <div style="height: 1px; background: rgba(255,255,255,0.05); margin-bottom: 1rem;"></div>
            <div style="font-family: 'DM Mono', monospace; font-size: 0.5rem; letter-spacing: 0.06em;
                        color: #2A3040; line-height: 1.8;">
                OILCORP ENERGIA LIMITED<br>v2.0 · NPA Reporting Suite
            </div>
        </div>
        """, unsafe_allow_html=True)

    return page


def main():
    page = render_sidebar()
    if page == "Order Request Report":
        page_order_report()
    elif page == "Daily Orders":
        page_daily_orders()


main()